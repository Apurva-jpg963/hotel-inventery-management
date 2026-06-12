import sys
import os
import threading
import webbrowser
from datetime import datetime, date
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

# ── Path setup ──────────────────────────────────────────────────────────────
def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS          # PyInstaller EXE
    except AttributeError:
        base_path = os.path.dirname(os.path.abspath(__file__))  # Normal run
    return os.path.join(base_path, relative_path)

BASE_DIR       = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_DIR   = os.path.join(BASE_DIR, 'templates')
STATIC_DIR     = os.path.join(BASE_DIR, 'static')
INVENTORY_FILE = os.path.join(BASE_DIR, 'inventory.xlsx')
FINANCE_FILE   = os.path.join(BASE_DIR, 'finance.xlsx')

app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=STATIC_DIR)
app.secret_key = 'hotel_saiprasad_secret_2024'

# ── Excel helpers ────────────────────────────────────────────────────────────
INVENTORY_HEADERS = [
    'Item ID', 'Item Name', 'Category', 'Quantity', 'Unit',
    'Supplier Name', 'Purchase Date', 'Minimum Stock Level', 'Status', 'Last Updated'
]
FINANCE_HEADERS = [
    'Transaction ID', 'Date', 'Type', 'Category', 'Amount', 'Description'
]

def init_excel(filepath, headers):
    if not os.path.exists(filepath):
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1a3a5c')
            cell.alignment = Alignment(horizontal='center')
        wb.save(filepath)

def load_workbook_data(filepath):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = rows[0]
    data = []
    for row in rows[1:]:
        if any(cell is not None for cell in row):
            data.append(dict(zip(headers, row)))
    return data

def save_all_rows(filepath, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sheet1'
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a3a5c')
        cell.alignment = Alignment(horizontal='center')
    for r_idx, row in enumerate(rows, 2):
        for col, h in enumerate(headers, 1):
            ws.cell(row=r_idx, column=col, value=row.get(h, ''))
    wb.save(filepath)

def next_inv_id():
    data = load_workbook_data(INVENTORY_FILE)
    if not data:
        return 'INV-0001'
    nums = []
    for d in data:
        try:
            nums.append(int(str(d.get('Item ID', '')).split('-')[1]))
        except:
            pass
    return f'INV-{(max(nums)+1 if nums else 1):04d}'

def next_fin_id():
    data = load_workbook_data(FINANCE_FILE)
    if not data:
        return 'TXN-0001'
    nums = []
    for d in data:
        try:
            nums.append(int(str(d.get('Transaction ID', '')).split('-')[1]))
        except:
            pass
    return f'TXN-{(max(nums)+1 if nums else 1):04d}'

def calc_status(qty, min_stock):
    try:
        qty       = int(qty)       if qty       is not None else 0
        min_stock = int(min_stock) if min_stock is not None else 0
    except:
        qty, min_stock = 0, 0
    if qty == 0:
        return 'Out of Stock'
    elif qty <= min_stock:
        return 'Low Stock'
    return 'Available'

# ── Init files ───────────────────────────────────────────────────────────────
init_excel(INVENTORY_FILE, INVENTORY_HEADERS)
init_excel(FINANCE_FILE,   FINANCE_HEADERS)

# ── Dashboard ────────────────────────────────────────────────────────────────
@app.route('/')
def dashboard():
    inv_data  = load_workbook_data(INVENTORY_FILE)
    fin_data  = load_workbook_data(FINANCE_FILE)
    today_str     = str(date.today())
    current_month = datetime.today().strftime('%Y-%m')

    total_items  = len(inv_data)
    low_stock    = sum(1 for i in inv_data if str(i.get('Status', '')) == 'Low Stock')
    out_of_stock = sum(1 for i in inv_data if str(i.get('Status', '')) == 'Out of Stock')
    categories   = len(set(i.get('Category', '') for i in inv_data if i.get('Category')))

    today_rev = today_exp = monthly_rev = monthly_exp = 0.0
    for t in fin_data:
        d = str(t.get('Date', ''))[:10]
        try:
            amt = float(t.get('Amount', 0) or 0)
        except:
            amt = 0.0
        if t.get('Type') == 'Revenue':
            if d == today_str:               today_rev   += amt
            if d.startswith(current_month):  monthly_rev += amt
        elif t.get('Type') == 'Expense':
            if d == today_str:               today_exp   += amt
            if d.startswith(current_month):  monthly_exp += amt

    today_profit   = today_rev   - today_exp
    monthly_profit = monthly_rev - monthly_exp

    recent_inv = sorted(inv_data, key=lambda x: str(x.get('Last Updated', '')), reverse=True)[:5]
    recent_fin = sorted(fin_data, key=lambda x: str(x.get('Date', '')),         reverse=True)[:5]
    low_alerts = [i for i in inv_data if str(i.get('Status', '')) == 'Low Stock'][:5]
    oos_alerts = [i for i in inv_data if str(i.get('Status', '')) == 'Out of Stock'][:5]

    return render_template('dashboard.html',
        total_items=total_items, low_stock=low_stock,
        out_of_stock=out_of_stock, categories=categories,
        today_rev=today_rev, today_exp=today_exp,
        today_profit=today_profit, monthly_profit=monthly_profit,
        monthly_rev=monthly_rev, monthly_exp=monthly_exp,
        recent_inv=recent_inv, recent_fin=recent_fin,
        low_alerts=low_alerts, oos_alerts=oos_alerts)

# ── Inventory ────────────────────────────────────────────────────────────────
@app.route('/inventory')
def inventory():
    data   = load_workbook_data(INVENTORY_FILE)
    search = request.args.get('search', '').strip().lower()
    cat_f  = request.args.get('category', '')
    stat_f = request.args.get('status', '')

    if search:
        data = [d for d in data if
                search in str(d.get('Item Name',     '')).lower() or
                search in str(d.get('Supplier Name', '')).lower() or
                search in str(d.get('Item ID',       '')).lower()]
    if cat_f:  data = [d for d in data if d.get('Category') == cat_f]
    if stat_f: data = [d for d in data if d.get('Status')   == stat_f]

    categories = ['Housekeeping', 'Food & Beverage', 'Room Supplies', 'Maintenance', 'Office Supplies']
    return render_template('inventory.html', items=data, categories=categories,
                           search=search, cat_f=cat_f, stat_f=stat_f)

@app.route('/inventory/add', methods=['GET', 'POST'])
def add_inventory():
    categories = ['Housekeeping', 'Food & Beverage', 'Room Supplies', 'Maintenance', 'Office Supplies']
    if request.method == 'POST':
        qty       = request.form.get('quantity',  '0')
        min_stock = request.form.get('min_stock', '0')
        new_item  = {
            'Item ID':             next_inv_id(),
            'Item Name':           request.form.get('item_name', '').strip(),
            'Category':            request.form.get('category',  ''),
            'Quantity':            qty,
            'Unit':                request.form.get('unit',      '').strip(),
            'Supplier Name':       request.form.get('supplier',  '').strip(),
            'Purchase Date':       request.form.get('purchase_date', ''),
            'Minimum Stock Level': min_stock,
            'Status':              calc_status(qty, min_stock),
            'Last Updated':        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
        if not new_item['Item Name']:
            flash('Item name is required.', 'danger')
            return render_template('inventory_form.html', item=None, categories=categories, action='Add')
        data = load_workbook_data(INVENTORY_FILE)
        data.append(new_item)
        save_all_rows(INVENTORY_FILE, INVENTORY_HEADERS, data)
        flash(f"Item '{new_item['Item Name']}' added successfully!", 'success')
        return redirect(url_for('inventory'))
    return render_template('inventory_form.html', item=None, categories=categories, action='Add')

@app.route('/inventory/edit/<item_id>', methods=['GET', 'POST'])
def edit_inventory(item_id):
    categories = ['Housekeeping', 'Food & Beverage', 'Room Supplies', 'Maintenance', 'Office Supplies']
    data = load_workbook_data(INVENTORY_FILE)
    item = next((d for d in data if str(d.get('Item ID')) == item_id), None)
    if not item:
        flash('Item not found.', 'danger')
        return redirect(url_for('inventory'))
    if request.method == 'POST':
        qty       = request.form.get('quantity',  '0')
        min_stock = request.form.get('min_stock', '0')
        for d in data:
            if str(d.get('Item ID')) == item_id:
                d['Item Name']           = request.form.get('item_name', '').strip()
                d['Category']            = request.form.get('category',  '')
                d['Quantity']            = qty
                d['Unit']                = request.form.get('unit',      '').strip()
                d['Supplier Name']       = request.form.get('supplier',  '').strip()
                d['Purchase Date']       = request.form.get('purchase_date', '')
                d['Minimum Stock Level'] = min_stock
                d['Status']              = calc_status(qty, min_stock)
                d['Last Updated']        = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        save_all_rows(INVENTORY_FILE, INVENTORY_HEADERS, data)
        flash('Item updated successfully!', 'success')
        return redirect(url_for('inventory'))
    return render_template('inventory_form.html', item=item, categories=categories, action='Edit')

@app.route('/inventory/delete/<item_id>', methods=['POST'])
def delete_inventory(item_id):
    data = [d for d in load_workbook_data(INVENTORY_FILE) if str(d.get('Item ID')) != item_id]
    save_all_rows(INVENTORY_FILE, INVENTORY_HEADERS, data)
    flash('Item deleted successfully.', 'success')
    return redirect(url_for('inventory'))

@app.route('/inventory/export')
def export_inventory():
    data = load_workbook_data(INVENTORY_FILE)
    wb = Workbook(); ws = wb.active; ws.title = 'Inventory'
    for col, h in enumerate(INVENTORY_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a3a5c')
    for r_idx, row in enumerate(data, 2):
        for col, h in enumerate(INVENTORY_HEADERS, 1):
            ws.cell(row=r_idx, column=col, value=row.get(h, ''))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='inventory_export.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Finance ──────────────────────────────────────────────────────────────────
@app.route('/finance')
def finance():
    data   = load_workbook_data(FINANCE_FILE)
    search = request.args.get('search',   '').strip().lower()
    type_f = request.args.get('type',     '')
    cat_f  = request.args.get('category', '')
    date_f = request.args.get('date',     '')

    if search:
        data = [d for d in data if
                search in str(d.get('Transaction ID', '')).lower() or
                search in str(d.get('Description',    '')).lower()]
    if type_f:  data = [d for d in data if d.get('Type')     == type_f]
    if cat_f:   data = [d for d in data if d.get('Category') == cat_f]
    if date_f:  data = [d for d in data if str(d.get('Date', ''))[:10] == date_f]

    today_str     = str(date.today())
    current_month = datetime.today().strftime('%Y-%m')
    today_rev = today_exp = monthly_rev = monthly_exp = 0.0
    for t in load_workbook_data(FINANCE_FILE):
        d = str(t.get('Date', ''))[:10]
        try:
            amt = float(t.get('Amount', 0) or 0)
        except:
            amt = 0.0
        if t.get('Type') == 'Revenue':
            if d == today_str:              today_rev   += amt
            if d.startswith(current_month): monthly_rev += amt
        elif t.get('Type') == 'Expense':
            if d == today_str:              today_exp   += amt
            if d.startswith(current_month): monthly_exp += amt

    rev_cats = ['Room Booking', 'Restaurant', 'Event Booking', 'Other Income']
    exp_cats = ['Maintenance', 'Housekeeping', 'Food Purchase', 'Electricity',
                'Water', 'Staff Salary', 'Miscellaneous']

    return render_template('finance.html', transactions=data,
        today_rev=today_rev, today_exp=today_exp,
        today_profit=today_rev - today_exp,
        monthly_rev=monthly_rev, monthly_exp=monthly_exp,
        monthly_profit=monthly_rev - monthly_exp,
        rev_cats=rev_cats, exp_cats=exp_cats,
        all_cats=rev_cats + exp_cats,
        search=search, type_f=type_f, cat_f=cat_f, date_f=date_f)

@app.route('/finance/add', methods=['GET', 'POST'])
def add_finance():
    rev_cats = ['Room Booking', 'Restaurant', 'Event Booking', 'Other Income']
    exp_cats = ['Maintenance', 'Housekeeping', 'Food Purchase', 'Electricity',
                'Water', 'Staff Salary', 'Miscellaneous']
    if request.method == 'POST':
        amt = request.form.get('amount', '0')
        try:
            if float(amt) <= 0:
                raise ValueError
        except:
            flash('Amount must be greater than 0.', 'danger')
            return render_template('finance_form.html', txn=None,
                                   rev_cats=rev_cats, exp_cats=exp_cats, action='Add')
        txn = {
            'Transaction ID': next_fin_id(),
            'Date':        request.form.get('date', str(date.today())),
            'Type':        request.form.get('txn_type', 'Revenue'),
            'Category':    request.form.get('category', ''),
            'Amount':      amt,
            'Description': request.form.get('description', '').strip(),
        }
        data = load_workbook_data(FINANCE_FILE)
        data.append(txn)
        save_all_rows(FINANCE_FILE, FINANCE_HEADERS, data)
        flash('Transaction added successfully!', 'success')
        return redirect(url_for('finance'))
    return render_template('finance_form.html', txn=None,
                           rev_cats=rev_cats, exp_cats=exp_cats, action='Add')

@app.route('/finance/edit/<txn_id>', methods=['GET', 'POST'])
def edit_finance(txn_id):
    rev_cats = ['Room Booking', 'Restaurant', 'Event Booking', 'Other Income']
    exp_cats = ['Maintenance', 'Housekeeping', 'Food Purchase', 'Electricity',
                'Water', 'Staff Salary', 'Miscellaneous']
    data = load_workbook_data(FINANCE_FILE)
    txn  = next((d for d in data if str(d.get('Transaction ID')) == txn_id), None)
    if not txn:
        flash('Transaction not found.', 'danger')
        return redirect(url_for('finance'))
    if request.method == 'POST':
        for d in data:
            if str(d.get('Transaction ID')) == txn_id:
                d['Date']        = request.form.get('date', str(date.today()))
                d['Type']        = request.form.get('txn_type', 'Revenue')
                d['Category']    = request.form.get('category', '')
                d['Amount']      = request.form.get('amount', '0')
                d['Description'] = request.form.get('description', '').strip()
        save_all_rows(FINANCE_FILE, FINANCE_HEADERS, data)
        flash('Transaction updated successfully!', 'success')
        return redirect(url_for('finance'))
    return render_template('finance_form.html', txn=txn,
                           rev_cats=rev_cats, exp_cats=exp_cats, action='Edit')

@app.route('/finance/delete/<txn_id>', methods=['POST'])
def delete_finance(txn_id):
    data = [d for d in load_workbook_data(FINANCE_FILE)
            if str(d.get('Transaction ID')) != txn_id]
    save_all_rows(FINANCE_FILE, FINANCE_HEADERS, data)
    flash('Transaction deleted successfully.', 'success')
    return redirect(url_for('finance'))

@app.route('/finance/export')
def export_finance():
    data = load_workbook_data(FINANCE_FILE)
    wb = Workbook(); ws = wb.active; ws.title = 'Finance'
    for col, h in enumerate(FINANCE_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1a3a5c')
    for r_idx, row in enumerate(data, 2):
        for col, h in enumerate(FINANCE_HEADERS, 1):
            ws.cell(row=r_idx, column=col, value=row.get(h, ''))
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, download_name='finance_export.xlsx', as_attachment=True,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

# ── Run ───────────────────────────────────────────────────────────────────────
def open_browser():
    webbrowser.open('http://127.0.0.1:5000')

if __name__ == '__main__':
    threading.Timer(1.2, open_browser).start()
    app.run(host='127.0.0.1', port=5000, debug=False, use_reloader=False)
